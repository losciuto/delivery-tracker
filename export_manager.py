"""
Export/Import functionality for Delivery Tracker.
Supports CSV, JSON, and Excel (.xlsx/.xls) formats.
"""
import csv
import json
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path
import utils

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    utils.logger.warning("openpyxl not available, Excel export disabled")

try:
    import xlrd
    XLS_AVAILABLE = True
except ImportError:
    XLS_AVAILABLE = False
    utils.logger.warning("xlrd not available, .xls import disabled")

logger = utils.get_logger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# Column mapping: DB field -> list of accepted header names (case-insensitive)
# ─────────────────────────────────────────────────────────────────────────────
COLUMN_MAPPING: Dict[str, List[str]] = {
    'order_date':         ['data ordine', 'order_date', 'data', "data dell'ordine", 'date'],
    'platform':           ['piattaforma', 'platform', 'store', 'negozio'],
    'description':        ['descrizione', 'description', 'articolo', 'nome prodotto', 'product name', 'item'],
    'seller':             ['venditore', 'seller', 'fornitore'],
    'destination':        ['destinazione', 'destination', 'indirizzo'],
    'link':               ['link', 'url', 'collegamento'],
    'quantity':           ['quantità', 'q.tà', 'quantity', 'qty', 'quantita'],
    'estimated_delivery': ['cons. prevista', 'estimated_delivery', 'consegna', 'consegna prevista', 'consegna stimata', 'delivery date'],
    'alarm_enabled':      ['alarm_enabled', 'allarme'],
    'is_delivered':       ['consegnato', 'is_delivered', 'delivered'],
    'position':           ['posizione', 'position'],
    'notes':              ['note', 'notes', 'annotazioni'],
    'category':           ['categoria', 'category'],
    'tracking_number':    ['n. tracking', 'tracking_number', 'tracking', 'numero tracking', 'numero di tracking', 'codice tracking'],
    'carrier':            ['vettore', 'carrier', 'corriere'],
    'last_mile_carrier':  ['ultimo miglio', 'last_mile_carrier', 'last mile'],
    'site_order_id':      ['id ordine sito', 'site_order_id', 'id ordine', 'order id', 'numero ordine'],
    'status':             ['stato', 'status'],
    'price':              ['prezzo prodotto (eur)', 'price', 'prezzo', 'costo', 'prezzo prodotto', 'prezzo unitario'],
    'image_url':          ['immagine', 'image_url', 'image url', 'foto', 'url immagine'],
}

# Reverse lookup: normalized header -> DB field
_HEADER_TO_FIELD: Dict[str, str] = {}
for _field, _aliases in COLUMN_MAPPING.items():
    for _alias in _aliases:
        _HEADER_TO_FIELD[_alias.lower().strip()] = _field


def _normalize_header(h: str) -> str:
    """Normalize a column header for lookup."""
    return str(h).lower().strip()


def _build_header_map(headers: List[str]) -> Dict[str, int]:
    """
    Given a list of raw column headers, return a dict mapping
    DB field name -> column index. Unknown columns are ignored.
    """
    mapping = {}
    for idx, raw in enumerate(headers):
        normalized = _normalize_header(raw)
        field = _HEADER_TO_FIELD.get(normalized)
        if field and field not in mapping:
            mapping[field] = idx
    return mapping


def _cast_row(row_dict: Dict[str, Any]) -> Dict[str, Any]:
    """Cast/clean values from a raw imported row."""
    result = dict(row_dict)

    # Quantity: int
    if 'quantity' in result:
        try:
            result['quantity'] = int(float(str(result['quantity']))) if result['quantity'] not in ('', None) else 1
        except (ValueError, TypeError):
            result['quantity'] = 1

    # Price: float or None
    if 'price' in result:
        try:
            val = str(result['price']).replace(',', '.').replace('€', '').strip()
            result['price'] = float(val) if val else None
        except (ValueError, TypeError):
            result['price'] = None

    # is_delivered: bool
    if 'is_delivered' in result:
        v = str(result['is_delivered']).lower().strip()
        result['is_delivered'] = v in ('true', '1', 'yes', 'sì', 'si', 'consegnato')

    # alarm_enabled: bool
    if 'alarm_enabled' in result:
        v = str(result['alarm_enabled']).lower().strip()
        result['alarm_enabled'] = v in ('true', '1', 'yes', 'sì', 'si')

    # Dates: normalize to ISO string or None
    for date_field in ('order_date', 'estimated_delivery'):
        if date_field in result:
            val = result[date_field]
            if val:
                parsed = utils.DateHelper.parse_smart(str(val))
                result[date_field] = parsed.strftime('%Y-%m-%d') if parsed else str(val)

    # Strip strings
    for k, v in result.items():
        if isinstance(v, str):
            result[k] = v.strip()

    return result


def detect_import_format(filepath: str) -> Optional[str]:
    """Detect the import format from file extension."""
    ext = Path(filepath).suffix.lower()
    if ext == '.xlsx':
        return 'xlsx'
    elif ext == '.xls':
        return 'xls'
    elif ext == '.csv':
        return 'csv'
    elif ext == '.json':
        return 'json'
    return None




class ExportManager:
    """Manager for exporting orders to various formats"""
    
    @staticmethod
    def export_to_excel(filepath: str, orders: List[Dict[str, Any]]) -> bool:
        """Export orders to Excel with formatting"""
        if not EXCEL_AVAILABLE:
            logger.error("Excel export not available: openpyxl not installed")
            return False
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Ordini"
            
            # Define headers
            headers = [
                "ID", "Data Ordine", "Piattaforma", "Venditore", "Destinazione",
                "Descrizione", "Link", "Quantità", "Consegna Prevista",
                "Posizione", "Categoria", "Consegnato", "Note"
            ]
            
            # Header style
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Data style
            border = Border(
                left=Side(style='thin', color='E0E0E0'),
                right=Side(style='thin', color='E0E0E0'),
                top=Side(style='thin', color='E0E0E0'),
                bottom=Side(style='thin', color='E0E0E0')
            )
            
            # Write data
            for row_num, order in enumerate(orders, 2):
                # Determine row color based on status
                if order.get('is_delivered'):
                    fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                elif order.get('estimated_delivery'):
                    est_date = utils.DateHelper.parse_date(order['estimated_delivery'])
                    if est_date:
                        status = utils.DateHelper.get_date_status(est_date)
                        if status == 'overdue':
                            fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                        elif status == 'due_today':
                            fill = PatternFill(start_color="FFE082", end_color="FFE082", fill_type="solid")
                        elif status == 'upcoming':
                            fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                        else:
                            fill = None
                    else:
                        fill = None
                else:
                    fill = None
                
                # Write cells
                data = [
                    order.get('id', ''),
                    order.get('order_date', ''),
                    order.get('platform', ''),
                    order.get('seller', ''),
                    order.get('destination', ''),
                    order.get('description', ''),
                    order.get('link', ''),
                    order.get('quantity', 1),
                    order.get('estimated_delivery', ''),
                    order.get('position', ''),
                    order.get('category', ''),
                    "Sì" if order.get('is_delivered') else "No",
                    order.get('notes', '')
                ]
                
                for col_num, value in enumerate(data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    if fill:
                        cell.fill = fill
                    
                    # Add hyperlink for link column
                    if col_num == 7 and value:  # Link column
                        cell.hyperlink = value
                        cell.font = Font(color="0000FF", underline="single")
            
            # Auto-adjust column widths
            for col_num, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_num)
                
                # Calculate max width
                max_length = len(header)
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
                    for cell in row:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            ws.freeze_panes = "A2"
            
            # Save workbook
            wb.save(filepath)
            logger.info(f"Exported {len(orders)} orders to Excel: {filepath}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            return False
    
    @staticmethod
    def export_to_csv(filepath: str, orders: List[Dict[str, Any]]) -> bool:
        """Export orders to CSV file"""
        try:
            if not orders:
                return True
            
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = [
                    'id', 'order_date', 'platform', 'seller', 'destination',
                    'description', 'link', 'quantity', 'estimated_delivery',
                    'position', 'category', 'is_delivered', 'notes'
                ]
                
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames, extrasaction='ignore')
                writer.writeheader()
                writer.writerows(orders)
            
            logger.info(f"Exported {len(orders)} orders to CSV: {filepath}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            return False
    
    @staticmethod
    def export_to_json(filepath: str, orders: List[Dict[str, Any]]) -> bool:
        """Export orders to JSON file"""
        try:
            with open(filepath, 'w', encoding='utf-8') as jsonfile:
                json.dump(orders, jsonfile, indent=2, ensure_ascii=False)
            
            logger.info(f"Exported {len(orders)} orders to JSON: {filepath}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting to JSON: {e}")
            return False
    
    @staticmethod
    def import_from_json(filepath: str) -> Tuple[bool, List[Dict[str, Any]]]:
        """Import orders from JSON file (exported by this app or compatible format)."""
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)

            if isinstance(data, dict):
                # Maybe wrapped: {"orders": [...]}
                data = data.get('orders', [])

            cleaned = []
            for order in data:
                order.pop('id', None)
                order.pop('created_at', None)
                order.pop('updated_at', None)
                cleaned.append(_cast_row(order))

            logger.info(f"Loaded {len(cleaned)} orders from JSON: {filepath}")
            return True, cleaned

        except Exception as e:
            logger.error(f"Error importing from JSON: {e}")
            return False, []

    @staticmethod
    def import_from_csv(filepath: str) -> Tuple[bool, List[Dict[str, Any]]]:
        """Import orders from CSV with dynamic column mapping and encoding fallback."""
        orders = []
        for encoding in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252'):
            try:
                with open(filepath, newline='', encoding=encoding) as f:
                    reader = csv.reader(f)
                    headers = next(reader, None)
                    if not headers:
                        return True, []

                    header_map = _build_header_map(headers)

                    for raw_row in reader:
                        if not any(c.strip() for c in raw_row):
                            continue  # skip empty lines
                        row_dict = {}
                        for field, col_idx in header_map.items():
                            if col_idx < len(raw_row):
                                row_dict[field] = raw_row[col_idx]
                        if row_dict:
                            orders.append(_cast_row(row_dict))

                logger.info(f"Loaded {len(orders)} orders from CSV ({encoding}): {filepath}")
                return True, orders

            except UnicodeDecodeError:
                continue
            except Exception as e:
                logger.error(f"Error importing from CSV: {e}")
                return False, []

        logger.error(f"Could not decode CSV file: {filepath}")
        return False, []

    @staticmethod
    def import_from_excel(filepath: str) -> Tuple[bool, List[Dict[str, Any]]]:
        """Import orders from Excel (.xlsx or .xls) with dynamic column mapping."""
        ext = Path(filepath).suffix.lower()
        try:
            orders = []

            if ext == '.xlsx':
                if not EXCEL_AVAILABLE:
                    return False, []
                wb = load_workbook(filepath, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                wb.close()

                if not rows:
                    return True, []

                headers = [str(h) if h is not None else '' for h in rows[0]]
                header_map = _build_header_map(headers)

                for raw_row in rows[1:]:
                    if not any(v for v in raw_row if v is not None):
                        continue
                    row_dict = {}
                    for field, col_idx in header_map.items():
                        if col_idx < len(raw_row):
                            val = raw_row[col_idx]
                            # openpyxl may return datetime objects
                            if hasattr(val, 'strftime'):
                                val = val.strftime('%Y-%m-%d')
                            row_dict[field] = val if val is not None else ''
                    if row_dict:
                        orders.append(_cast_row(row_dict))

            elif ext == '.xls':
                if not XLS_AVAILABLE:
                    logger.error("xlrd not available for .xls import")
                    return False, []
                book = xlrd.open_workbook(filepath)
                sheet = book.sheet_by_index(0)

                if sheet.nrows == 0:
                    return True, []

                headers = [str(cell.value) for cell in sheet.row(0)]
                header_map = _build_header_map(headers)

                for row_idx in range(1, sheet.nrows):
                    raw_row = sheet.row(row_idx)
                    if not any(cell.value for cell in raw_row):
                        continue
                    row_dict = {}
                    for field, col_idx in header_map.items():
                        if col_idx < len(raw_row):
                            val = raw_row[col_idx].value
                            # xlrd may return floats for dates; convert if needed
                            if raw_row[col_idx].ctype == xlrd.XL_CELL_DATE:
                                try:
                                    dt = xlrd.xldate_as_datetime(val, book.datemode)
                                    val = dt.strftime('%Y-%m-%d')
                                except Exception:
                                    pass
                            row_dict[field] = val if val != '' else ''
                    if row_dict:
                        orders.append(_cast_row(row_dict))
            else:
                logger.error(f"Unsupported Excel format: {ext}")
                return False, []

            logger.info(f"Loaded {len(orders)} orders from Excel ({ext}): {filepath}")
            return True, orders

        except Exception as e:
            logger.error(f"Error importing from Excel: {e}")
            return False, []

    @staticmethod
    def import_auto(filepath: str) -> Tuple[bool, List[Dict[str, Any]], str]:
        """
        Auto-detect format and import.
        Returns (success, orders, error_message).
        """
        fmt = detect_import_format(filepath)
        if fmt is None:
            return False, [], "Formato file non supportato. Usa .xlsx, .xls, .csv o .json"

        if fmt in ('xlsx', 'xls'):
            ok, orders = ExportManager.import_from_excel(filepath)
        elif fmt == 'csv':
            ok, orders = ExportManager.import_from_csv(filepath)
        elif fmt == 'json':
            ok, orders = ExportManager.import_from_json(filepath)
        else:
            return False, [], "Formato non riconosciuto"

        if not ok:
            return False, [], "Errore durante la lettura del file"
        return True, orders, ""
